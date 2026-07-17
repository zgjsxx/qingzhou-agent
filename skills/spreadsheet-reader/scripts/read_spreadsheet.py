from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path


def truncate(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars]}\n... ({len(text) - max_chars} more characters)"


def format_rows(rows: list[list[str]], max_cols: int) -> list[str]:
    lines: list[str] = []
    for row in rows:
        cells = [str(cell or "").replace("\n", " ").strip() for cell in row[:max_cols]]
        if len(row) > max_cols:
            cells.append(f"... ({len(row) - max_cols} more columns)")
        lines.append(" | ".join(cells))
    return lines


def read_delimited(path: Path, delimiter: str, max_rows: int, max_cols: int) -> tuple[list[str], int]:
    rows: list[list[str]] = []
    row_count = 0
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row in reader:
            row_count += 1
            if row_count <= max_rows:
                rows.append(row)
    return format_rows(rows, max_cols), row_count


def read_xlsx(path: Path, max_sheets: int, max_rows: int, max_cols: int) -> list[str]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is not installed. Run: pip install -r requirements.txt") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        lines = [f"sheets: {', '.join(workbook.sheetnames)}"]
        for sheet_index, sheet_name in enumerate(workbook.sheetnames[:max_sheets], start=1):
            sheet = workbook[sheet_name]
            lines.extend(["", f"[sheet {sheet_index}] {sheet.title}"])
            lines.append(f"dimensions: rows={sheet.max_row}, columns={sheet.max_column}")
            rows = [
                ["" if value is None else str(value) for value in row]
                for row in sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True)
            ]
            lines.extend(format_rows(rows, max_cols))
            if sheet.max_row > max_rows:
                lines.append(f"... ({sheet.max_row - max_rows} more rows)")
            if sheet.max_column > max_cols:
                lines.append(f"... ({sheet.max_column - max_cols} more columns)")
        if len(workbook.sheetnames) > max_sheets:
            lines.append(f"... ({len(workbook.sheetnames) - max_sheets} more sheets)")
        return lines
    finally:
        workbook.close()


def read_xls(path: Path, max_sheets: int, max_rows: int, max_cols: int) -> list[str]:
    try:
        import xlrd
    except ImportError as exc:
        raise SystemExit("xlrd is not installed. Run: pip install -r requirements.txt") from exc

    workbook = xlrd.open_workbook(str(path))
    sheet_names = workbook.sheet_names()
    lines = [f"sheets: {', '.join(sheet_names)}"]
    for sheet_index, sheet_name in enumerate(sheet_names[:max_sheets], start=1):
        sheet = workbook.sheet_by_name(sheet_name)
        lines.extend(["", f"[sheet {sheet_index}] {sheet.name}"])
        lines.append(f"dimensions: rows={sheet.nrows}, columns={sheet.ncols}")
        rows = [
            [str(sheet.cell_value(row_index, col_index)) for col_index in range(min(sheet.ncols, max_cols))]
            for row_index in range(min(sheet.nrows, max_rows))
        ]
        lines.extend(format_rows(rows, max_cols))
        if sheet.nrows > max_rows:
            lines.append(f"... ({sheet.nrows - max_rows} more rows)")
        if sheet.ncols > max_cols:
            lines.append(f"... ({sheet.ncols - max_cols} more columns)")
    if len(sheet_names) > max_sheets:
        lines.append(f"... ({len(sheet_names) - max_sheets} more sheets)")
    return lines


def main() -> int:
    parser = argparse.ArgumentParser(description="Preview CSV, TSV, XLSX, and XLS spreadsheets.")
    parser.add_argument("path", help="Path to the spreadsheet.")
    parser.add_argument("--max-sheets", type=int, default=5)
    parser.add_argument("--max-rows", type=int, default=30)
    parser.add_argument("--max-cols", type=int, default=12)
    parser.add_argument("--max-chars", type=int, default=12000)
    parser.add_argument("--json", action="store_true", help="Emit JSON with metadata and text.")
    args = parser.parse_args()

    path = Path(args.path).expanduser().resolve()
    if not path.exists() or not path.is_file():
        raise SystemExit(f"file does not exist: {path}")

    suffix = path.suffix.lower()
    lines = ["[Spreadsheet]", f"path: {path}", f"type: {suffix.lstrip('.') or 'unknown'}"]
    if suffix == ".csv":
        preview, row_count = read_delimited(path, ",", args.max_rows, args.max_cols)
        lines.extend([f"rows: {row_count}", "", "[preview]", *preview])
        if row_count > args.max_rows:
            lines.append(f"... ({row_count - args.max_rows} more rows)")
    elif suffix == ".tsv":
        preview, row_count = read_delimited(path, "\t", args.max_rows, args.max_cols)
        lines.extend([f"rows: {row_count}", "", "[preview]", *preview])
        if row_count > args.max_rows:
            lines.append(f"... ({row_count - args.max_rows} more rows)")
    elif suffix == ".xlsx":
        lines.extend(read_xlsx(path, args.max_sheets, args.max_rows, args.max_cols))
    elif suffix == ".xls":
        lines.extend(read_xls(path, args.max_sheets, args.max_rows, args.max_cols))
    else:
        raise SystemExit("unsupported spreadsheet type; expected .csv, .tsv, .xlsx, or .xls")

    text = truncate("\n".join(lines), args.max_chars)
    if args.json:
        print(json.dumps({"path": str(path), "kind": suffix.lstrip("."), "text": text}, ensure_ascii=False, indent=2))
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
